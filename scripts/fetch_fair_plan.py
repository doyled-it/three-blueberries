"""Regenerate lib/data/ca-insurance.ts from primary sources.

California home insurance is the most volatile line in the payment and the one
most likely to be catastrophically wrong for a specific buyer. It is also the
one nobody publishes a clean county average for: the Department of Insurance
runs an interactive premium tool rather than a dataset, and the quote
aggregators that do publish city tables will not say where their numbers came
from.

What IS published, quarterly, by the body that writes the policies, is the
California FAIR Plan's own book: policy counts and written premium for every ZIP
code in the state, split by wildfire risk score and policy category. That is the
insurer of last resort, so it is not the price of a normal policy. It is
something arguably more useful, the price and the likelihood of the outcome that
wrecks a budget.

Denominator is the Department of Finance E-5, single detached units by county,
so "share of homes on the FAIR Plan" compares detached houses with detached
houses.

    uv run python scripts/fetch_fair_plan.py

Refuses to write the file unless the parse reconciles to the totals printed in
the source PDFs and all 58 counties are present. A silent partial parse here
would understate the risk in exactly the counties where it matters most.
"""

from __future__ import annotations

import io
import re
import sys
import urllib.request
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
import pypdf

# The FAIR Plan renames these every quarter with the as-of date in the filename.
# Find the current ones at https://www.cfpnet.com/key-statistics-data/
AS_OF = "2026-06-30"
POLICY_URL = "https://www.cfpnet.com/wp-content/uploads/2026/07/Policies-by-category-DWE-as-of-260630-DL-260717v001.pdf"
PREMIUM_URL = "https://www.cfpnet.com/wp-content/uploads/2026/07/Premium-by-category-DWE-as-of-260630-DL-260717v001.pdf"

# Department of Finance E-5, published every May.
E5_URL = (
    "https://dof.ca.gov/media/docs/forecasting/Demographics/estimates/"
    "e-5-population-and-housing-estimates-for-cities-counties-and-the-state-2020-2026/"
    "E-5_2026_InternetVersion.xlsx"
)
E5_AS_OF = "2026-01-01"

# Printed on the "All All All All" row of each PDF. The parse has to reproduce
# these exactly or we are silently dropping ZIP codes.
EXPECTED_POLICIES = 149_958 + 153_758 + 216_034
EXPECTED_PREMIUM = 213_704_028 + 481_981_444 + 863_313_547

COUNTY_COUNT = 58

# Each row carries five policy categories repeated across three wildfire risk
# bands: low, medium, high. Owner-occupied single family is the first column of
# each band, which is the only one a homebuyer is buying.
OWNER_OCCUPIED_COLUMNS = (0, 5, 10)
HIGH_RISK_COLUMN = 10
COLUMNS_PER_ROW = 15

ROW = re.compile(r"^(\d{5})\s+([A-Za-z][A-Za-z .'\-]*?)\s+([01])\s+([A-Za-z][A-Za-z ]*?)\s+(?=[\d,\-])(.*)$")

OUT = Path(__file__).resolve().parent.parent / "lib" / "data" / "ca-insurance.ts"


@dataclass
class County:
    policies: int = 0
    premium: int = 0
    high_risk_policies: int = 0
    high_risk_premium: int = 0
    zips: set[str] = field(default_factory=set)


def fetch(url: str) -> bytes:
    print(f"  fetching {url.rsplit('/', 1)[-1]}")
    with urllib.request.urlopen(url) as response:  # noqa: S310 - fixed, published URLs
        return response.read()


def parse_rows(pdf_bytes: bytes, *, money: bool) -> list[tuple[str, str, list[int]]]:
    """One tuple per ZIP code: (zip, county, 15 values)."""
    reader = pypdf.PdfReader(io.BytesIO(pdf_bytes))
    rows: list[tuple[str, str, list[int]]] = []
    unparsed = 0

    for page in reader.pages:
        for line in page.extract_text().split("\n"):
            line = line.strip()
            if not re.match(r"^\d{5}\s", line):
                continue
            match = ROW.match(line)
            if not match:
                unparsed += 1
                continue
            zip_code, county, _distressed, _region, rest = match.groups()
            if money:
                # "1,234$" for a value, a bare "-$" for zero.
                values = [0 if t == "-" else int(t.replace(",", "")) for t in re.findall(r"(-|[\d,]+)\$", rest)]
            else:
                values = [int(t.replace(",", "")) for t in re.findall(r"[\d,]+", rest)]
            if len(values) != COLUMNS_PER_ROW:
                unparsed += 1
                continue
            rows.append((zip_code, county.strip(), values))

    if unparsed:
        sys.exit(f"FAIL: {unparsed} ZIP rows did not parse. The layout has changed; fix the parser, do not ship.")
    return rows


def parse_detached_units(xlsx: bytes) -> dict[str, int]:
    """Single detached housing units by county, from the newest E-5 sheet."""
    workbook = openpyxl.load_workbook(io.BytesIO(xlsx), read_only=True, data_only=True)
    sheets = [s for s in workbook.sheetnames if s.startswith("E5CountyState")]
    if not sheets:
        sys.exit("FAIL: no E5CountyState sheet in the workbook.")
    sheet = workbook[sheets[-1]]
    print(f"  using sheet {sheets[-1]}")

    units: dict[str, int] = {}
    for row in sheet.iter_rows(values_only=True):
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()
        try:
            units[name] = int(row[5])
        except (TypeError, ValueError, IndexError):
            continue
    return units


def main() -> None:
    print("California FAIR Plan, residential policies by ZIP")
    policy_rows = parse_rows(fetch(POLICY_URL), money=False)
    premium_rows = parse_rows(fetch(PREMIUM_URL), money=True)

    premiums = {(z, c): v for z, c, v in premium_rows}
    if len(premiums) != len(policy_rows):
        sys.exit("FAIL: the two reports cover different ZIP codes.")

    counties: dict[str, County] = defaultdict(County)
    for zip_code, county, policy_values in policy_rows:
        premium_values = premiums.get((zip_code, county))
        if premium_values is None:
            sys.exit(f"FAIL: {zip_code} ({county}) is in the policy report but not the premium report.")
        entry = counties[county]
        entry.zips.add(zip_code)
        entry.policies += sum(policy_values[i] for i in OWNER_OCCUPIED_COLUMNS)
        entry.premium += sum(premium_values[i] for i in OWNER_OCCUPIED_COLUMNS)
        entry.high_risk_policies += policy_values[HIGH_RISK_COLUMN]
        entry.high_risk_premium += premium_values[HIGH_RISK_COLUMN]

    total_policies = sum(c.policies for c in counties.values())
    total_premium = sum(c.premium for c in counties.values())
    if total_policies != EXPECTED_POLICIES:
        sys.exit(f"FAIL: parsed {total_policies:,} policies, the report says {EXPECTED_POLICIES:,}.")
    if total_premium != EXPECTED_PREMIUM:
        sys.exit(f"FAIL: parsed ${total_premium:,} of premium, the report says ${EXPECTED_PREMIUM:,}.")
    if len(counties) != COUNTY_COUNT:
        sys.exit(f"FAIL: {len(counties)} counties, California has {COUNTY_COUNT}.")
    print(f"  reconciled: {total_policies:,} policies, ${total_premium:,} of premium, {len(counties)} counties")

    print("Department of Finance E-5, housing units by county")
    units = parse_detached_units(fetch(E5_URL))
    missing = sorted(c for c in counties if c not in units)
    if missing:
        sys.exit(f"FAIL: no housing unit count for {', '.join(missing)}.")

    statewide_average = round(total_premium / total_policies)

    lines = [
        "/**",
        " * GENERATED FILE. Do not hand-edit. Run `npm run data:insurance`.",
        " *",
        " * California FAIR Plan owner-occupied single-family residential policies and",
        f" * written premium by county, as of {AS_OF}, aggregated from the FAIR Plan's own",
        " * per-ZIP quarterly reports at https://www.cfpnet.com/key-statistics-data/",
        " *",
        " * `detachedUnits` is the Department of Finance E-5 count of single detached",
        f" * housing units in the county as of {E5_AS_OF}, so `share` compares detached",
        " * houses with detached houses.",
        " *",
        " * THIS IS NOT THE PRICE OF A NORMAL POLICY. The FAIR Plan is the insurer of",
        " * last resort: fire-only cover, usually needing a separate DIC policy on top",
        " * for everything else a homeowners policy does. It is what you pay when the",
        " * admitted market will not write you, which is why the share matters as much",
        " * as the premium.",
        " */",
        "",
        "import type { CaCounty } from \"./ca-loan-limits.ts\";",
        "",
        "export interface FairPlanCounty {",
        "  /** Owner-occupied single-family policies in force. */",
        "  policies: number;",
        "  /** Mean annual written premium per policy, dollars. */",
        "  averagePremium: number;",
        "  /** Mean premium in the county's high-wildfire-risk ZIP codes, null if it has none. */",
        "  highRiskPremium: number | null;",
        "  /** Single detached housing units in the county. */",
        "  detachedUnits: number;",
        "  /** Policies divided by detached units. */",
        "  share: number;",
        "}",
        "",
        f'export const FAIR_PLAN_AS_OF = "{AS_OF}";',
        f'export const FAIR_PLAN_UNITS_AS_OF = "{E5_AS_OF}";',
        "/** Mean owner-occupied single-family FAIR Plan premium across the state. */",
        f"export const FAIR_PLAN_STATEWIDE_AVERAGE = {statewide_average};",
        "",
        "export const FAIR_PLAN_BY_COUNTY: Record<CaCounty, FairPlanCounty> = {",
    ]

    for name in sorted(counties):
        entry = counties[name]
        detached = units[name]
        average = round(entry.premium / entry.policies) if entry.policies else 0
        high = round(entry.high_risk_premium / entry.high_risk_policies) if entry.high_risk_policies else None
        share = round(entry.policies / detached, 5) if detached else 0
        key = f'"{name}"' if " " in name else name
        lines.append(
            f"  {key}: {{ policies: {entry.policies}, averagePremium: {average}, "
            f"highRiskPremium: {high if high is not None else 'null'}, "
            f"detachedUnits: {detached}, share: {share} }},"
        )

    lines.append("};")
    lines.append("")

    OUT.write_text("\n".join(lines), encoding="utf-8")
    print(f"wrote {OUT.relative_to(Path.cwd())}")
    print(f"  statewide mean FAIR Plan premium ${statewide_average:,}")
    ranked = sorted(counties.items(), key=lambda kv: kv[1].policies / max(units[kv[0]], 1), reverse=True)
    for name, entry in ranked[:3]:
        share = entry.policies / units[name]
        print(f"  {name}: {share:.1%} of detached homes, ${round(entry.premium / entry.policies):,} average")


if __name__ == "__main__":
    main()
