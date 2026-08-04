"""Regenerates lib/data/history.ts: price history for all 58 California counties.

    npm run data:history

WHY THIS EXISTS IN THIS SHAPE

The site lets you pick any of 58 counties, and for a long time every one of them
was shown San Diego's price history with a note apologising for it. That is
backwards: the selector promises a county-specific answer, so the history has to
be county-specific too.

S&P CoreLogic Case-Shiller, which this replaced, publishes three California
metros. Permission to use it would not have fixed the problem, because the other
55 counties are simply not in it. FHFA publishes 28 California metros quarterly
and all 58 counties annually, as a US government work with no redistribution
restriction, so it is both the only source that can honour the selector and the
only one that can be committed to a public repository without asking anybody.

FOUR SOURCES, JOINED HERE

  FHFA HPI, metro, quarterly     the price series for 37 counties
  FHFA HPI, county, annual       the price series for the other 21
  Census/OMB delineation file    which county belongs to which metro
  Zillow ZHVI, county, monthly   the dollar anchor, since FHFA publishes an
                                 index with no units
  Freddie Mac PMMS via FRED      the rate that went with each period

THE SPLICE, AND THE SEAM IN IT

FHFA's expanded-data flavour is the accurate one: it combines conforming records
with county recorder and CoreLogic data, so it sees jumbo and cash sales, and it
measures the 2008 crash at -41.9% against Case-Shiller's -42%. But it only starts
in 1991, which truncates the 1990 peak and misses the early-1980s entirely.

All-transactions runs from 1975 but counts refinance APPRAISALS, which lag a
turning market: it reads that same crash as -35% bottoming two years late.

So each metro series is expanded-data from 1991, with all-transactions chain
linked on before it, scaled to meet at the join. That is standard index practice
and it buys the 1981 downturn and the true 1990 peak. It is also a SEAM: the two
families disagree on year-over-year growth by about 2.5 points on average inside
their overlap. The generated file says where the seam is, and `spliceMonth` is
exported so the UI can too.
"""

from __future__ import annotations

import csv
import io
import sys
import urllib.request
from collections import defaultdict
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
OUT = REPO / "lib" / "data" / "history.ts"

FHFA_MASTER = "https://www.fhfa.gov/hpi/download/monthly/hpi_master.csv"
FHFA_COUNTY = "https://www.fhfa.gov/hpi/download/annual/hpi_at_county.xlsx"
CBSA_DELINEATION = (
    "https://www2.census.gov/programs-surveys/metro-micro/geographies/"
    "reference-files/2023/delineation-files/list1_2023.xlsx"
)
ZHVI_COUNTY_SFR = (
    "https://files.zillowstatic.com/research/public_csvs/zhvi/"
    "County_zhvi_uc_sfr_tier_0.33_0.67_sm_sa_month.csv"
)
RATES = "https://fred.stlouisfed.org/graph/fredgraph.csv?id=MORTGAGE30US"

QUARTER_END = {1: "03", 2: "06", 3: "09", 4: "12"}

CA_COUNTIES = [
    "Alameda", "Alpine", "Amador", "Butte", "Calaveras", "Colusa", "Contra Costa", "Del Norte",
    "El Dorado", "Fresno", "Glenn", "Humboldt", "Imperial", "Inyo", "Kern", "Kings", "Lake",
    "Lassen", "Los Angeles", "Madera", "Marin", "Mariposa", "Mendocino", "Merced", "Modoc", "Mono",
    "Monterey", "Napa", "Nevada", "Orange", "Placer", "Plumas", "Riverside", "Sacramento",
    "San Benito", "San Bernardino", "San Diego", "San Francisco", "San Joaquin", "San Luis Obispo",
    "San Mateo", "Santa Barbara", "Santa Clara", "Santa Cruz", "Shasta", "Sierra", "Siskiyou",
    "Solano", "Sonoma", "Stanislaus", "Sutter", "Tehama", "Trinity", "Tulare", "Tuolumne",
    "Ventura", "Yolo", "Yuba",
]


def fetch(url: str, label: str) -> bytes:
    print(f"  fetching {label}")
    with urllib.request.urlopen(url) as response:  # noqa: S310 - fixed, published URLs
        return response.read()


def die(message: str) -> None:
    sys.exit(f"FAIL: {message}")


# ---------------------------------------------------------------------------
# Sources
# ---------------------------------------------------------------------------


METRO_NAMES: dict[str, str] = {}


def metro_series() -> dict[str, list[tuple[str, float]]]:
    """Chained quarterly index per FHFA metro place id."""
    text = fetch(FHFA_MASTER, "FHFA metro master").decode("utf-8-sig")
    expanded: dict[str, dict[str, float]] = defaultdict(dict)
    alltrans: dict[str, dict[str, float]] = defaultdict(dict)

    for row in csv.DictReader(io.StringIO(text)):
        if row["level"] != "MSA" or row["frequency"] != "quarterly":
            continue
        # FHFA's own name for the metro. Deriving a label from the first county
        # that maps to it labelled the Sacramento metro "El Dorado area", because
        # El Dorado sorts first among the four counties in it.
        METRO_NAMES.setdefault(row["place_id"], row["place_name"])
        month = f"{row['yr']}-{QUARTER_END[int(row['period'])]}"
        if row["hpi_flavor"] == "expanded-data" and row["index_sa"]:
            expanded[row["place_id"]][month] = float(row["index_sa"])
        elif row["hpi_flavor"] == "all-transactions" and row["index_nsa"]:
            alltrans[row["place_id"]][month] = float(row["index_nsa"])

    out: dict[str, list[tuple[str, float]]] = {}
    for place, recent in expanded.items():
        older = alltrans.get(place, {})
        join = min(recent)
        merged = dict(recent)
        # Chain-link: scale the older family so the two meet exactly at the join.
        if older and join in older and older[join] > 0:
            factor = recent[join] / older[join]
            for month, value in older.items():
                if month < join:
                    merged[month] = value * factor
        out[place] = sorted(merged.items())
    return out


def county_annual() -> dict[str, list[tuple[str, float]]]:
    """Annual index per California county, for the counties with no metro."""
    workbook = openpyxl.load_workbook(io.BytesIO(fetch(FHFA_COUNTY, "FHFA county annual")), read_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    out: dict[str, list[tuple[str, float]]] = defaultdict(list)
    started = False
    for row in sheet.iter_rows(values_only=True):
        if row and row[0] == "State":
            started = True
            continue
        if not started or not row or row[0] != "CA":
            continue
        county = str(row[1]).replace(" County", "").strip()
        try:
            value = float(row[5])
        except (TypeError, ValueError):
            continue
        # Stamp December, so the point sits at the end of the year it describes
        # and month arithmetic against the quarterly series stays comparable.
        out[county].append((f"{int(row[3])}-12", value))
    return {k: sorted(v) for k, v in out.items()}


def county_to_metro(available: set[str]) -> dict[str, str | None]:
    """County -> FHFA place id, from the OMB delineation rather than by name."""
    workbook = openpyxl.load_workbook(io.BytesIO(fetch(CBSA_DELINEATION, "Census delineation")), read_only=True)
    rows = list(workbook[workbook.sheetnames[0]].iter_rows(values_only=True))

    header = None
    start = 0
    for i, row in enumerate(rows[:8]):
        if row and any(c and "CBSA Code" in str(c) for c in row):
            header = [str(c).strip() if c else "" for c in row]
            start = i + 1
            break
    if header is None:
        die("no header row in the Census delineation file")
    col = {name: i for i, name in enumerate(header)}

    mapping: dict[str, str | None] = {county: None for county in CA_COUNTIES}
    for row in rows[start:]:
        if not row or not row[col["CBSA Code"]] or str(row[col["State Name"]]) != "California":
            continue
        county = str(row[col["County/County Equivalent"]]).replace(" County", "").strip()
        if county not in mapping:
            continue
        # FHFA keys on the metropolitan DIVISION where one exists, which is how
        # Los Angeles, Oakland, Anaheim, San Francisco and San Rafael are listed.
        division = row[col["Metropolitan Division Code"]]
        for candidate in (division, row[col["CBSA Code"]]):
            if candidate and str(int(candidate)) in available:
                mapping[county] = str(int(candidate))
                break
    return mapping


def anchors() -> dict[str, int]:
    """Latest Zillow ZHVI single-family typical value, per county."""
    text = fetch(ZHVI_COUNTY_SFR, "Zillow ZHVI county").decode("utf-8-sig")
    rows = list(csv.DictReader(io.StringIO(text)))
    months = [k for k in rows[0] if k and k[0].isdigit()]

    out: dict[str, int] = {}
    for row in rows:
        if row.get("State") != "CA":
            continue
        county = str(row["RegionName"]).replace(" County", "").strip()
        latest = next((row[m] for m in reversed(months) if row[m]), None)
        if latest:
            out[county] = round(float(latest))
    return out


def rates() -> list[tuple[str, float]]:
    text = fetch(RATES, "Freddie Mac PMMS").decode("utf-8-sig")
    out = []
    for line in text.strip().split("\n")[1:]:
        date, value = line.split(",")[:2]
        if value and value != ".":
            out.append((date, float(value)))
    return out


# ---------------------------------------------------------------------------
# Assembly
# ---------------------------------------------------------------------------


def main() -> None:
    print("California price history, all 58 counties")
    metros = metro_series()
    counties = county_annual()
    mapping = county_to_metro(set(metros))
    anchor = anchors()
    weekly = rates()

    missing_anchor = [c for c in CA_COUNTIES if c not in anchor]
    if missing_anchor:
        die(f"no Zillow anchor for {', '.join(missing_anchor)}")

    quarterly_counties = [c for c in CA_COUNTIES if mapping[c]]
    annual_counties = [c for c in CA_COUNTIES if not mapping[c]]
    unresolved = [c for c in annual_counties if c not in counties]
    if unresolved:
        die(f"no metro AND no county series for {', '.join(unresolved)}")
    print(f"  {len(quarterly_counties)} counties on a quarterly metro series")
    print(f"  {len(annual_counties)} counties on their own annual series")

    # Rates, averaged to each frequency so the two sides of every row describe
    # the same span of time.
    quarterly_rates: dict[str, list[float]] = defaultdict(list)
    annual_rates: dict[str, list[float]] = defaultdict(list)
    for date, value in weekly:
        year, month = date.split("-")[:2]
        quarterly_rates[f"{year}-{QUARTER_END[(int(month) - 1) // 3 + 1]}"].append(value)
        annual_rates[f"{year}-12"].append(value)

    def join(series: list[tuple[str, float]], buckets: dict[str, list[float]]) -> list[tuple[str, float, float]]:
        rows = []
        for month, index in series:
            bucket = buckets.get(month)
            if bucket:
                rows.append((month, round(index, 2), round(sum(bucket) / len(bucket), 3)))
        return rows

    used_metros = sorted({mapping[c] for c in quarterly_counties if mapping[c]})
    metro_rows = {place: join(metros[place], quarterly_rates) for place in used_metros}
    county_rows = {c: join(counties[c], annual_rates) for c in annual_counties}

    for place, rows in metro_rows.items():
        if len(rows) < 130:
            die(f"metro {place} has only {len(rows)} quarters")
    for county, rows in county_rows.items():
        if len(rows) < 30:
            die(f"county {county} has only {len(rows)} annual points")

    splice = {}
    for place in used_metros:
        expanded_start = min(m for m, _ in metros[place] if m >= "1991-01")
        splice[place] = expanded_start

    def fmt(rows: list[tuple[str, float, float]]) -> str:
        return ",".join(f'["{m}",{i},{r}]' for m, i, r in rows)

    sd = metro_rows[mapping["San Diego"]]
    lines = [
        "// GENERATED FILE, do not edit by hand.",
        "// Regenerate with: npm run data:history",
        "//",
        "// Price history for every California county, so the county selector at the top",
        "// of the form means something on the history panels too. See",
        "// scripts/fetch_history.py for why this is FHFA rather than Case-Shiller, and",
        "// where the 1991 seam in each metro series comes from.",
        "//",
        "// Sources, none of which require anyone's permission to redistribute:",
        "//   FHFA House Price Index, metro quarterly and county annual (US government work)",
        "//   Census/OMB delineation file, for county to metro (US government work)",
        "//   Zillow ZHVI single-family, for the dollar anchor (free, attribution required)",
        "//   Freddie Mac PMMS via FRED, for the rate",
        "",
        'import type { CaCounty } from "./ca-loan-limits.ts";',
        "",
        "/** [YYYY-MM at the END of the period, price index, average 30-year rate percent] */",
        "export type HistoryRow = readonly [string, number, number];",
        "",
        "export interface CountyHistory {",
        "  /** The series itself, oldest first. */",
        "  rows: readonly HistoryRow[];",
        "  /** Months between consecutive rows: 3 for a metro, 12 for a county. */",
        "  stepMonths: number;",
        "  /** What the series actually measures, for the UI to name. */",
        "  place: string;",
        "  /** Where the chained older index gives way to the accurate newer one. */",
        "  spliceMonth: string | null;",
        "  /** Zillow ZHVI single-family typical value, the dollar anchor. */",
        "  anchorPrice: number;",
        "}",
        "",
    ]

    lines.append("const METRO_ROWS: Record<string, readonly HistoryRow[]> = {")
    for place in used_metros:
        lines.append(f'  "{place}": [{fmt(metro_rows[place])}],')
    lines.append("};")
    lines.append("")

    lines.append("const COUNTY_ROWS: Record<string, readonly HistoryRow[]> = {")
    for county in annual_counties:
        lines.append(f'  "{county}": [{fmt(county_rows[county])}],')
    lines.append("};")
    lines.append("")

    lines.append("const METRO_NAMES: Record<string, string> = {")
    for place in used_metros:
        # Trim FHFA's trailing state and division markers: the site is only ever
        # showing California, and "(MSAD)" means nothing to a reader.
        name = METRO_NAMES[place].replace(" (MSAD)", "").replace(", CA", "")
        lines.append(f'  "{place}": "{name}",')
    lines.append("};")
    lines.append("")

    lines.append("const SPLICE: Record<string, string> = {")
    for place in used_metros:
        lines.append(f'  "{place}": "{splice[place]}",')
    lines.append("};")
    lines.append("")

    lines.append("/** County to FHFA metro place id, or null where the county has no metro series. */")
    lines.append("export const COUNTY_METRO: Record<CaCounty, string | null> = {")
    for county in CA_COUNTIES:
        place = mapping[county]
        key = f'"{county}"' if " " in county else county
        value = f'"{place}"' if place else "null"
        lines.append(f"  {key}: {value},")
    lines.append("};")
    lines.append("")

    lines.append("/** Zillow ZHVI single-family typical value by county, the dollar anchor. */")
    lines.append("export const COUNTY_ANCHOR: Record<CaCounty, number> = {")
    for county in CA_COUNTIES:
        key = f'"{county}"' if " " in county else county
        lines.append(f"  {key}: {anchor[county]},")
    lines.append("};")
    lines.append("")

    lines.append("export function historyFor(county: CaCounty): CountyHistory {")
    lines.append("  const place = COUNTY_METRO[county];")
    lines.append("  if (place) {")
    lines.append("    return {")
    lines.append("      rows: METRO_ROWS[place]!,")
    lines.append("      stepMonths: 3,")
    lines.append("      place: METRO_NAMES[place]!,")
    lines.append("      spliceMonth: SPLICE[place] ?? null,")
    lines.append("      anchorPrice: COUNTY_ANCHOR[county],")
    lines.append("    };")
    lines.append("  }")
    lines.append("  return {")
    lines.append("    rows: COUNTY_ROWS[county]!,")
    lines.append("    stepMonths: 12,")
    lines.append("    place: `${county} County`,")
    lines.append("    spliceMonth: null,")
    lines.append("    anchorPrice: COUNTY_ANCHOR[county],")
    lines.append("  };")
    lines.append("}")
    lines.append("")
    lines.append("/** The default the site opens on. */")
    lines.append('export const DEFAULT_COUNTY = "San Diego" as const;')
    lines.append(f'export const HISTORY_LATEST_MONTH = "{sd[-1][0]}";')
    lines.append("")

    OUT.write_text("\n".join(lines), encoding="utf-8")
    size = OUT.stat().st_size // 1024
    print(f"wrote {OUT.relative_to(REPO)} ({size} KB)")
    print(f"  San Diego: {len(sd)} quarters, {sd[0][0]} to {sd[-1][0]}, anchor ${anchor['San Diego']:,}")
    print(f"  {len(used_metros)} distinct metro series, {len(annual_counties)} annual counties")


if __name__ == "__main__":
    main()
